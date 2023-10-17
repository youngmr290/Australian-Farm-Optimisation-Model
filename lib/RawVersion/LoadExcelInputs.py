"""

These are inputs that are expected to remain constant between regions and properties, this includes:

* Prices of inputs
* Value of outputs (grain, wool ,meat)
* Interest rates & depreciation rates
* Machinery options
* Sheep parameters and definition of genotypes

author: young
"""

##python modules
import pickle as pkl
import pandas as pd
import numpy as np
import os.path
import sys
from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange

from . import LoadExp as exp
from ..AfoLogic import StructuralInputs as sinp
from ..AfoLogic import UniversalInputs as uinp
from ..AfoLogic import PropertyInputs as pinp
from ..AfoLogic import relativeFile


def f_load_fs(fs_use_pkl, fs_use_number):
    pkl_fs = {}
    if fs_use_pkl:
        print(f'pkl_fs{fs_use_number} being used.')
        pkl_fs_path = relativeFile.find(__file__, "../../pkl", f"pkl_fs{fs_use_number}.pkl")
        with open(pkl_fs_path,"rb") as f:
            pkl_fs = pkl.load(f)
    return pkl_fs

def f_load_stubble():
    ##read in category info frpm xl
    ###build path
    cat_propn_s1_ks2 = pd.read_excel(relativeFile.findExcel("stubble sim.xlsx"),header=None, engine='openpyxl')
    return cat_propn_s1_ks2

def f_load_phases():
    '''
    Load existing rotation info. If it doesnt exist then pass.

    The rotations are checked later to see if they match the users needs for the given trial (ie full vs pinp).
    The rotations are also masked later. See pinp.f1_phases()

    '''
    ##rotation phases - read in from excel
    rot_xl_path = relativeFile.findExcel("Rotation.xlsx")
    try:
        phases_r = pd.read_excel(rot_xl_path, sheet_name='rotation list', header=None, index_col=0,
                                 engine='openpyxl').T.reset_index(drop=True).T  # reset the col headers to std ie 0,1,2 etc
        rot_req = pd.read_excel(rot_xl_path, sheet_name='rotation_req', header=None,
                                engine='openpyxl')  # , index_col = [0,1]) #couldn't get it to read in with multiindex for some reason
        rot_prov = pd.read_excel(rot_xl_path, sheet_name='rotation_prov', header=None,
                                 engine='openpyxl')  # , index_col = [0,1]) #couldn't get it to read in with multiindex for some reason
        s_rotcon1 = pd.read_excel(rot_xl_path, sheet_name='rotation con1 set', header=None, index_col=0, engine='openpyxl')

    except FileNotFoundError:
        phases_r = None
        rot_req = None
        rot_prov = None
        s_rotcon1 = None

    return {"phases_r": phases_r, "rot_req": rot_req, "rot_prov": rot_prov, "s_rotcon1": s_rotcon1}

def f_load_excel_default_inputs(load_all_pinp=False, trial_pinp=None):
    '''Function to load inputs from excel (univeral, structural, property, price variation, rotation and stubble)'''

    #########################################################################################################################################################################################################
    #########################################################################################################################################################################################################
    # read in excel - structural
    #########################################################################################################################################################################################################
    #########################################################################################################################################################################################################

    ##dict to store structural inputs
    sinp_defaults={}

    structural_xl_path = relativeFile.findExcel("Structural.xlsx")

    ##read from excel
    print('Reading structural inputs from Excel',end=' ',flush=True)
    ##general
    sinp_defaults['general_inp'] = xl_all_named_ranges(structural_xl_path,"General")

    ##sheep inputs
    sinp_defaults['stock_inp'] = xl_all_named_ranges(structural_xl_path,'Stock',numpy=True)

    ##sa inputs (these variables can have sensitivity applied from exp.xl
    sinp_defaults['structuralsa_inp'] = xl_all_named_ranges(structural_xl_path,'StructuralSA',numpy=True)

    sinp_defaults['rep_inp'] = xl_all_named_ranges(structural_xl_path,"Report Settings")

    print('- finished')


    #########################################################################################################################################################################################################
    #########################################################################################################################################################################################################
    #read in excel - property
    #########################################################################################################################################################################################################
    #########################################################################################################################################################################################################

    ##determine which properties are used in current exp
    if not load_all_pinp:
        pinp_defaults_req = trial_pinp.dropna().unique()
    else: #this is required for the web app which needs to load all excel.
        pinp_defaults_req = ['GSW', 'CWW', 'SWV']

    ##read in inputs
    pinp_defaults={}
    for property in pinp_defaults_req:
        pinp_defaults[property] = {}
        ##build path.
        property_xl_path = relativeFile.findExcel("Property_{0}.xlsx".format(property))

        ##read from excel
        print('Reading property {0} inputs from Excel'.format(property), end=' ', flush=True)
        pinp_defaults[property]['general_inp'] = xl_all_named_ranges(property_xl_path,"General", numpy=True)

        pinp_defaults[property]['labour_inp'] = xl_all_named_ranges(property_xl_path,"Labour")

        pinp_defaults[property]['crop_inp'] = xl_all_named_ranges(property_xl_path,"Crop")

        pinp_defaults[property]['cropgraze_inp'] = xl_all_named_ranges(property_xl_path,"CropGrazing", numpy=True)

        pinp_defaults[property]['saltbush_inp'] = xl_all_named_ranges(property_xl_path,"Saltbush", numpy=True)

        pinp_defaults[property]['mach_inp'] = xl_all_named_ranges(property_xl_path,"Mach")

        pinp_defaults[property]['stubble_inp'] = xl_all_named_ranges(property_xl_path,"CropResidue", numpy=True)

        pinp_defaults[property]['finance_inp'] = xl_all_named_ranges(property_xl_path,"Finance")

        pinp_defaults[property]['period_inp'] = xl_all_named_ranges(property_xl_path,"Periods", numpy=True) #automatically read in the periods as dates

        pinp_defaults[property]['sup_inp'] = xl_all_named_ranges(property_xl_path,"Sup Feed")

        pinp_defaults[property]['sheep_inp']  = xl_all_named_ranges(property_xl_path, 'Sheep', numpy=True)

        pinp_defaults[property]['feedsupply_inp']  = xl_all_named_ranges(property_xl_path, 'FeedSupply', numpy=True)

        pinp_defaults[property]['mvf_inp']  = xl_all_named_ranges(property_xl_path, 'MVEnergy', numpy=True)

        pinp_defaults[property]['pasture_inp']=dict()
        for pasture in sinp_defaults["general_inp"]['pastures'][pinp_defaults[property]['general_inp']['i_pastures_exist']]:
            pinp_defaults[property]['pasture_inp'][pasture] = xl_all_named_ranges(property_xl_path, pasture, numpy=True)

        print('- finished')


    #########################################################################################################################################################################################################
    #########################################################################################################################################################################################################
    #read in excel - universal
    #########################################################################################################################################################################################################
    #########################################################################################################################################################################################################

    ##dict to store universal inputs
    uinp_defaults={}

    ##build path
    universal_xl_path = relativeFile.findExcel("Universal.xlsx")

    ##read from excel
    print('Reading universal inputs from Excel', end=' ', flush=True)
    ##general
    uinp_defaults['general_inp'] = xl_all_named_ranges(universal_xl_path,"General")

    ##prices
    uinp_defaults['price_inp'] = xl_all_named_ranges(universal_xl_path,"Price")

    ##Finance inputs
    uinp_defaults['finance_inp'] = xl_all_named_ranges(universal_xl_path,"Finance")

    ##mach inputs - general
    uinp_defaults['mach_general_inp'] = xl_all_named_ranges(universal_xl_path,"Mach General")

    ##sup inputs
    uinp_defaults['sup_inp'] = xl_all_named_ranges(universal_xl_path,"Sup Feed")

    ##crop inputs
    uinp_defaults['crop_inp'] = xl_all_named_ranges(universal_xl_path,"Crop Sim")

    ##sheep inputs
    uinp_defaults['sheep_inp'] = xl_all_named_ranges(universal_xl_path, 'Sheep', numpy=True)
    uinp_defaults['parameters_inp'] = xl_all_named_ranges(universal_xl_path, 'Parameters', numpy=True)
    uinp_defaults['pastparameters_inp'] = xl_all_named_ranges(universal_xl_path, 'PastParameters', numpy=True)

    ##mach options
    ###create a dict to store all options - this allows the user to select an option
    uinp_defaults['machine_options_dict_inp']={}
    uinp_defaults['machine_options_dict_inp'][1] = xl_all_named_ranges(universal_xl_path,"Mach 1")
    uinp_defaults['machine_options_dict_inp'][2] = xl_all_named_ranges(universal_xl_path,"Mach 2")


    ##read in price variation inputs from xl - this might change
    price_variation_inp = {}
    ###build path
    pricescenarios_xl_path = relativeFile.findExcel("PriceScenarios.xlsx")
    ###read price info
    price_variation_inp['grain_price_scalar_c1z'] = pd.read_excel(pricescenarios_xl_path,sheet_name='grain',index_col=0,header=0,engine='openpyxl')
    price_variation_inp['meat_price_scalar_c1z'] = pd.read_excel(pricescenarios_xl_path,sheet_name='meat',index_col=0,header=0,engine='openpyxl').values
    price_variation_inp['wool_price_scalar_c1z'] = pd.read_excel(pricescenarios_xl_path,sheet_name='wool',index_col=0,header=0,engine='openpyxl').values
    price_variation_inp['prob_c1'] = pd.read_excel(pricescenarios_xl_path,sheet_name='prob',index_col=0,header=0,engine='openpyxl').squeeze()
    price_variation_inp['len_c1'] = len(price_variation_inp['prob_c1'])
    ###load into uinp_defaults dict
    uinp_defaults['price_variation_inp'] = price_variation_inp

    print('- finished')

    ##reshape multi dimension default inputs (they come from excel/database in 2d)
    sinp.f_reshape_sinp_defaults(sinp_defaults)
    uinp.f_reshape_uinp_defaults(uinp_defaults)
    pinp.f_reshape_pinp_defaults(pinp_defaults, sinp_defaults) #pass in sinp defaults because some of the axis lengths are required

    return sinp_defaults, uinp_defaults, pinp_defaults


# caching the excel spreadsheets to greatly speed things up
excel_cache = {}
def load_excel(filename):
    if filename in excel_cache:
        return excel_cache[filename]
    else:
        # print(f"[excel] loading: {filename}")
        sys.stdout.flush()
        excel_cache[filename] = load_workbook(filename, data_only=True, read_only=False)
        return excel_cache[filename]

# requires being passed the filename and sheetname for the workbook that will be accessed
# returns a dict with the key being the excel rangename
# the dict includes: numbers (where the rangename is a single cell), lists (where the rangename is one dimensional) and dataframes (where the range is 2 dimensional)
# If the range is 2D the function converts the first row to the dataframe column names and the first col to index names
# if you don't want this you can reset index using index.reset or something and probs the similar for cols
# Testing showed readonly = False was quicker than true. But still not as fast as pandas
# (may not exist anymore) now it causes problems sometimes locking you out of excel because it is readonly - closing doesn't fix issue (wb._archive.close())

def xl_all_named_ranges(filename, targetsheets, rangename=None, numpy=False,
                        datatype=None):  # read all range names defined in the list targetsheets and return a dictionary of lists or dataframes
    ''' Read data from named ranges in an Excel workbook.

    Parameters:
    filename is an Excel workbook name (including the extension).
    targetsheets is a list of (or a single) worksheet names from which to read the range names.
    rangename is an optional argument. If not included then all rangenames are read. If included only that name is read in.
    numpy is an optional boolean argument. If True it will assign the input array to a numpy
    datatype: you can use this parameter to select the data type of the numpy arrays. if a value doesn't match the dtype it gets a nan

    Returns:
    A dictionary that includes key that correspond to the rangenames
    '''

    wb = load_excel(filename)
    # t_wb = wb
    parameters = {}
    ## convert targetsheets to lowercase and handle both an individual name and a list
    try:
        targetsheets = targetsheets.lower()
    except:  # targetsheets is a list
        targetsheets = [name.lower() for name in targetsheets]

    for dn in wb.defined_names.definedName[:]:
        if rangename is None or dn.name == rangename:
            try:
                sheet_name, cell_range = list(dn.destinations)[
                    0]  # if it is a non-contiguous range dn.destinations would need to be looped through
                # print (dn.name, cell_range)
                if sheet_name.lower() in targetsheets:  # in to check list of sheet names
                    try:
                        cr = CellRange(cell_range)
                        width = cr.max_col - cr.min_col
                        length = cr.max_row - cr.min_row
                        ws = wb[sheet_name]
                        # print (dn.name, sheet_name, cell_range, length, width)
                        if not width and not length:  # the range is a single cell & is not iterable
                            parameters[dn.name] = ws[cell_range].value
                        elif not width:  # the range is only 1 column & is not iterable across the row
                            parameters[dn.name] = np.asarray(
                                [cell.value for cell in [row[0] for row in ws[cell_range]]], dtype=datatype)
                        elif not length:  # the range is 1 row & is iterable across columns
                            for row in ws[cell_range]:
                                parameters[dn.name] = np.asarray([cell.value for cell in row], dtype=datatype)
                        elif numpy == True:
                            parameters[dn.name] = np.asarray([[cell.value for cell in row] for row in ws[cell_range]],
                                                             dtype=datatype)
                        else:  # the range is a region & is iterable across rows and columns
                            df = pd.DataFrame([cell.value for cell in row] for row in ws[cell_range])
                            # df = pd.DataFrame(cells)
                            # print(df)
                            ##set headers
                            df.rename(columns=df.iloc[0], inplace=True)
                            ###drop row that had header names (renaming is more like a copy than a cut)
                            df.drop(df.index[0], inplace=True)
                            ##set index
                            # df = df.set_index(df.iloc[:, 0], append=True)
                            df = df.rename(index=df.iloc[:, 0]).rename_axis(df.iloc[:, 0].name)
                            ###drop the first col because renaming/set_index is more like copy than cut hence it doesn't make the index col one just rename index to match col one
                            df = df.drop(df.columns[[0]],
                                         axis=1)  # for some reason this will chuck an error in the index values are int and there is nothing in the top left cell of the df...seems like a bug in python
                            ## manipulate data into cheapest format - results in mainly float32 (strings are still objects) - without this each value is treated as an object (objects use up much more memory) - this change reduced fert df from 150mbs to 20mbs
                            parameters[dn.name] = df.apply(pd.to_numeric, errors='ignore', downcast='float')
                    except TypeError:
                        pass
            except IndexError:
                pass
    wb.close()
    return parameters  # t_wb #
