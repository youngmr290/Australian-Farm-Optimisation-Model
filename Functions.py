"""
Functions used across the model (multi module functions). These functions don't import other AFO modules
(hence don't initialise inputs inside the function). All function parameters are passed in when calling the function.

@author: young
"""
import pandas as pd
import numpy as np
import pickle as pkl
from dateutil import relativedelta as rdelta
import os.path
import glob
import pyomo.environ as pe
import sys
import copy

#this module shouldn't import other AFO modules
import Exceptions as exc #can import exceptions because exceptions imports no modules


na = np.newaxis
################################################
#function to read in excel named ranges to a df#
################################################
#requires being passed the filename and sheetname for the workbook that will be accessed
#returns a dict with the key being the excel rangename
#the dict includes: numbers (where the rangename is a single cell), lists (where the rangename is one dimensional) and dataframes (where the range is 2 dimensional)
#If the range is 2D the function converts the first row to the dataframe column names and the first col to index names
#if you don't want this you can reset index using index.reset or something and probs the similar for cols
#Testing showed readonly = False was quicker than true. But still not as fast as pandas
# (may not exist anymore) now it causes problems sometimes locking you out of excel because it is readonly - closing doesn't fix issue (wb._archive.close())

def xl_all_named_ranges(filename, targetsheets, rangename=None,numpy=False,datatype=None):     # read all range names defined in the list targetsheets and return a dictionary of lists or dataframes
    ''' Read data from named ranges in an Excel workbook.

    Parameters:
    filename is an Excel workbook name (including the extension).
    targetsheets is a list of (or a single) worksheet names from which to read the range names.
    rangename is an optional argument. If not included then all rangenames are read. If included only that name is read in.
    numpy is an optional boolean argument. If True it will assign the input array to a numpy
    datatype: you can use this parameter to select the data type of the numpy arrays. if a value doesn't match the dtype it gets a nan
    
    Returns:
    A dictionary that includes key that correspond to the rangenames
    '''
    from openpyxl import load_workbook
    from openpyxl.worksheet.cell_range import CellRange

    wb = load_workbook(filename, data_only=True, read_only=False)
    # t_wb = wb
    parameters = {}
    ## convert targetsheets to lowercase and handle both an individual name and a list
    try:
        targetsheets = targetsheets.lower()
    except:   #targetsheets is a list
        targetsheets = [name.lower() for name in targetsheets]

    for dn in wb.defined_names.definedName[:]:
        if rangename is None or dn.name == rangename:
            try:
                sheet_name, cell_range = list(dn.destinations)[0]        # if it is a non-contiguous range dn.destinations would need to be looped through
                #print (dn.name, cell_range)
                if sheet_name.lower() in targetsheets:     # in to check list of sheet names
                    try:
                        cr = CellRange(cell_range)
                        width = cr.max_col - cr.min_col
                        length = cr.max_row - cr.min_row
                        ws = wb[sheet_name]
                        #print (dn.name, sheet_name, cell_range, length, width)
                        if not width and not length:            # the range is a single cell & is not iterable
                            parameters[dn.name] = ws[cell_range].value
                        elif not width:                         # the range is only 1 column & is not iterable across the row
                            parameters[dn.name] = np.asarray([cell.value for cell in [row[0] for row in ws[cell_range]]],dtype=datatype)
                        elif not length:                        # the range is 1 row & is iterable across columns
                            for row in ws[cell_range]:
                                parameters[dn.name] = np.asarray([cell.value for cell in row],dtype=datatype)
                        elif numpy == True:
                            parameters[dn.name] = np.asarray([[cell.value for cell in row] for row in ws[cell_range]],dtype=datatype)
                        else:                                   # the range is a region & is iterable across rows and columns
                            df = pd.DataFrame([cell.value for cell in row] for row in ws[cell_range])
                            #df = pd.DataFrame(cells)
                            #print(df)
                            df.rename(columns=df.iloc[0],inplace=True)
                            ## drop row that had header names (renaming is more like a copy than a cut)
                            df.drop(df.index[0],inplace=True)
                            df = df.set_index(df.iloc[:,0]) #could use rename ie df.rename(index=df.iloc[:,0],inplace=True)
                            ## now have to drop the first col because renaming/set_index is more like copy than cut hence it doesn't make the index col one just rename index to match col one
                            df = df.drop(df.columns[[0]],axis=1) #for some reason this will chuck an error in the index values are int and there is nothing in the top left cell of the df...seems like a bug in python
                            ## manipulate data into cheapest format - results in mainly float32 (strings are still objects) - without this each value is treated as an object (objects use up much more memory) - this change reduced fert df from 150mbs to 20mbs
                            parameters[dn.name] = df.apply(pd.to_numeric, errors='ignore', downcast='float')
                    except TypeError:
                        pass
            except IndexError:
                pass
    wb.close()
    return parameters #t_wb #

def f_convert_to_inf(input):
    input=input.astype('object') #have to convert to object so that when the stuff below is assigned it is not assigned as a string
    ##convert -- to -inf
    mask = input=='--'
    input[mask]=-np.inf
    ##convert ++ to inf
    mask = input=='++'
    input[mask]=np.inf
    ##convert 'True' to True (string to bool) - because array is read in as string
    mask = input=='True'
    input[mask]=True
    ##convert 'False' to False (string to bool) - because array is read in as string
    mask = input=='False'
    input[mask]=False
    return input.astype('float')

###########################
#general functions        #
###########################

#this is the fastest function for building cartesian products. Doesn't make much diff for small ones but up to 50% faster for big ones
def cartesian_product_simple_transpose(arrays):
    la = len(arrays)
    try:
        dtype = np.result_type(*arrays)
        arr = np.empty([la] + [len(a) for a in arrays], dtype=dtype)
    except TypeError:
        arr = np.empty([la] + [len(a) for a in arrays], dtype='U25')
    for i, a in enumerate(np.ix_(*arrays)):
        arr[i, ...] = a
    return arr.reshape(la, -1).T


def searchsort_multiple_dim(a, v, axis_a0, axis_v0, axis_a1=None, axis_v1=None, side='left'):
    '''
    Find the indices into a sorted array 'a' such that, if the corresponding elements in 'v' were inserted before the indices, the order of 'a' would be preserved.
    It does this iteratively down the specified axis (therefore the specified axis must be present in both 'a' and 'v'

    Parameters:
        a: 3-D array_like
        Input array. Must be sorted in ascending order.

        v: array_like
        Values to insert into a.

        axis_a0: int
        The position of axis to iterate along. a1 & v1 axis must be same length
        axis_a1: int
        The position of axis to iterate along. a1 & v1 axis must be same length

    '''
    # final = np.zeros_like(v)
    # slc_v = [slice(None)] * len(v.shape)
    # slc_a = [slice(None)] * len(a.shape)
    # for i in range(v.shape[axis_v]):
    #     slc_v[axis_v] = slice(i, i+1)
    #     slc_a[axis_a] = slice(i, i+1)
    #     final[tuple(slc_v)] = np.searchsorted(np.squeeze(a[tuple(slc_a)]), v[tuple(slc_v)])
    final = np.zeros_like(v).astype(int)
    slc_a = [slice(None)] * len(a.shape)
    slc_v = [slice(None)] * len(v.shape)
    for i in range(a.shape[axis_a0]):
        if axis_a1 is not None:
            for j in range(a.shape[axis_a1]):
                slc_a[axis_a0] = slice(i, i+1)
                slc_a[axis_a1] = slice(j, j+1)
                slc_v[axis_v0] = slice(i, i+1)
                slc_v[axis_v1] = slice(j, j+1)
                final[tuple(slc_v)] = np.searchsorted(np.squeeze(a[tuple(slc_a)]), v[tuple(slc_v)], side)
        else:
            slc_a[axis_a0] = slice(i, i+1)
            slc_v[axis_v0] = slice(i, i+1)
            final[tuple(slc_v)] = np.searchsorted(np.squeeze(a[tuple(slc_a)]), v[tuple(slc_v)], side)
    return final

#print(timeit.timeit(phases2,number=100)/100)
#
def f_expand(array, left_pos=0, swap=False, ax1=0, ax2=1, right_pos=0, left_pos2=0, right_pos2=0
                     , left_pos3=0, right_pos3=0, condition = None, axis = 0, swap2=False, ax1_2=1, ax2_2=2,
                     condition2=None, axis2=0, condition3=None, axis3=0, left_pos4=0, right_pos4=0, move=False, source=0, dest=1):
    '''
    *note: if adding two sets of new axis add from right to left (then the pos variables align)
    *note: mask applied last (after expanding and reshaping)

    Parameters
    ----------
    array : array
        parameter array - input from excel.
    left_pos : int
        position of axis to the left of where the new axis will be added.
    swap : boolean, optional
        do you want to swap the first two axis?. The default is False.
    right_pos : int, optional
        the position of the axis to the right of the singleton axis being added. The default is -1, for when the axis to the right is g?.
    left_pos2 : int
        position of axis to the left of where the new axis will be added.
    right_pos2 : int, optional
        the position of the axis to the right of the singleton axis being added. The default is -1, for when the axis to the right is g?.
    condition: boolean, optional
        mask used to slice given axis.
    axis: int, optional
        axis to apply mask to.

    Returns
    -------
    expands, swaps axis if required and apply a mask to a given axis if required.
    '''
    ##convert int to 1d array if required
    if type(array) == int:
        array = np.array([array])
    ##swap axis if necessary
    if swap:
        array = np.swapaxes(array, ax1, ax2)
    ##swap axis if necessary
    if swap2:
        array = np.swapaxes(array, ax1_2, ax2_2)
    ##move axis if necessary
    if move:
        array = np.moveaxis(array, source=source, destination=dest)
    ##get axis into correct position 1
    if left_pos != 0:
        extra_axes = tuple(range((left_pos + 1), right_pos))
    else: extra_axes = ()
    array = np.expand_dims(array, axis = extra_axes)
    ##get axis into correct position 2 (some arrays need singleton axis added in multiple places ie separated by a used axis)
    if left_pos2 != 0:
        extra_axes = tuple(range((left_pos2 + 1), right_pos2))
    else: extra_axes = ()
    array = np.expand_dims(array, axis = extra_axes)
    ##get axis into correct position 3 (some arrays need singleton axis added in multiple places ie separated by a used axis)
    if left_pos3 != 0:
        extra_axes = tuple(range((left_pos3 + 1), right_pos3))
    else: extra_axes = ()
    array = np.expand_dims(array, axis = extra_axes)
    ##get axis into correct position 4 (some arrays need singleton axis added in multiple places ie separated by a used axis)
    if left_pos4 != 0:
        extra_axes = tuple(range((left_pos4 + 1), right_pos4))
    else: extra_axes = ()
    array = np.expand_dims(array, axis = extra_axes)
    ##apply mask if required
    if condition is not None: #see if condition exists
        if type(condition) == bool: #check if array or single value - note array of T & F is not type bool (it is array)
            condition= np.asarray([condition]) #convert to numpy if it is singular input
            array = np.compress(condition, array, axis)
        else:
            array = np.compress(condition, array, axis)
    ##apply mask if required
    if condition2 is not None: #see if condition exists
        if type(condition2) == bool: #check if array or single value - note array of T & F is not type bool (it is array)
            condition2= np.asarray([condition2]) #convert to numpy if it is singular input
            array = np.compress(condition2, array, axis2)
        else:
            array = np.compress(condition2, array, axis2)
    ##apply mask if required
    if condition3 is not None: #see if condition exists
        if type(condition3) == bool: #check if array or single value - note array of T & F is not type bool (it is array)
            condition3= np.asarray([condition3]) #convert to numpy if it is singular input
            array = np.compress(condition3, array, axis3)
        else:
            array = np.compress(condition3, array, axis3)
    return array

def f_update(existing_value, new_value, mask_for_new):
    '''
    Parameters
    ----------
    existing_value : numpy array or float or int
        values you want when mask = false.
    new_value : numpy array or float or int
        values you want when mask = true.
    mask_for_new : boolean mask
        boolean mask for the final axis of the array (typically the g axis).

    Returns
    -------
    Numpy array
        returns a combination of the two input arrays determined by the mask. Note: multiplying by true return the original number and multiplying by false results in 0.

    '''
    ##dtype for output (primarily needed for pp when int32 and float32 create float64 which we don't want)
    ##if the new value is an object (eg contains '-') then we want to return the original dtype otherwise return the biggest dtype

    if isinstance(new_value,np.ndarray) and isinstance(existing_value,np.ndarray):
        if new_value.dtype == object:
            dtype = existing_value.dtype
        else:
            dtype = max(existing_value.dtype, new_value.dtype)
    elif isinstance(new_value,np.ndarray):
        dtype = new_value.dtype
    elif isinstance(existing_value,np.ndarray):
        dtype = existing_value.dtype
    elif isinstance(mask_for_new,np.ndarray):
        pass #if both values are int/float and mask is numpy then just ignore dtype
    elif type(existing_value)==type(new_value):
        dtype = type(existing_value)
    elif isinstance(new_value, str): #if it is '-' (used in exp) then retain the original dtype
        dtype = type(existing_value)


    ##convert '-' to 0 (because '-' * False == '' which causes and error when you add to existing value)
    ##need a try and except in case the new value is not a numpy array (ie it is a single value)
    try:
        if np.any(new_value.astype('object')=='-'): #needs to be an object to preform elementwise comparison
                new_value[new_value=='-'] = 0
                new_value = new_value.astype(float) #need to convert to number because if str it chucks error below
    except AttributeError:
        if new_value=='-':
            new_value = 0
    #todo using a masked array may allow f_update to handle situation that have a nan value that is masked - MRY addition: i couldn't get this method to actually work
    # updated = np.ma.masked_array(existing_value, mask_for_new) + np.ma.maskedarray(new_value, np.logical_not(mask_for_new))  #used 'not' rather than '~' because ~False == -1 rather than True (not the case for np.arrays only if bool is single - as it is for sire in some situations)
    updated = existing_value * np.logical_not(mask_for_new) + new_value * mask_for_new #used not rather than ~ because ~False == -1 not True (not the case for np.arrays only if bool is single - as it is for sire in some situations)

    ##convert back to original dtype because adding float32 and int32 returns float64. And sometimes we don't want this eg postprocessing
    ###use try except because sometimes a single int is update eg in the first iteration on generator. this causes error because only numpy arrays have .dtype.
    try:
        updated = updated.astype(dtype)
    except AttributeError:
        pass
    except UnboundLocalError:
        pass
    ###used for core python dtype eg floats/int/str
    try:
        updated = dtype(updated)
    except TypeError:
        pass
    except UnboundLocalError:
        pass
    ###error check
    try:
        if updated.dtype == object:
            print('dtype error in f_update (type object is being returned)') #will give warning if ever returning a numpy object
    except: pass

    return updated


def f_weighted_average(array, weights, axis, keepdims=False, non_zero=False, den_weights=1):
    '''
    Calculates weighted average (similar to np.average however this will handle:
        if the sum of the weights is 0 (np.average doesn't handle this)
        keeping the axis (using the keepdims argument)
    'non-zero' handles how the average is calculated
    Note: if non-zero is false then when sum weights = 0 the numbers being averaged also = 0 (so can divide by 1 instead of 0)
    The function is also called from the reporting module with den_weights. den_weights can be 0, in which case 'non-zero' handles how the average is calculated
    axis averaged along can be retained - default it is dropped.


    :param array:
    :param weights:
    :param axis:
    :param keepdims:
    :param non_zero: how to handle a weight of 0. True returns the numerator, False (default) returns 0
    :param den_weights: array: broadcastable to weights. This is used to weight the denominator (used in reporting)
    :return:
    '''
    if non_zero:
        ##for some situations (production) if numbers are 0 we don't want to return 0 we want to return the original value
        weights=f_update(weights,1,np.all(weights==0, axis=axis, keepdims=True))
    weighted_array = np.sum(array * weights, axis=axis, keepdims=keepdims)
    weights = np.broadcast_to(np.sum(weights * den_weights, axis=axis, keepdims=keepdims), weighted_array.shape)
    # den_weight = np.broadcast_to(np.sum(den_weight, axis=axis, keepdims=keepdims), weighted_array.shape)
    averaged_array = np.zeros_like(weighted_array)
    mask = weights!=0
    averaged_array[mask] = weighted_array[mask] / weights[mask]
    return averaged_array

def f_divide(numerator, denominator, dtype='float64', option=0):
    '''
    Elementwise divides two arrays.
    If the denominator = 0 then return value depends on 'option'
     option == 0 then return 0
     option == 1 then return 1 if the numerator is also 0

     option == 1 will also return 1 if both the numerator and denominator are np.inf

    '''
    numerator, denominator = np.broadcast_arrays(numerator, denominator)
    result = np.zeros(numerator.shape, dtype=dtype) #make it a float in case the numerator is int
    ##use ~np.isclose to capture when the denominator is 0 within rounding tolerances
    mask = ~np.isclose(denominator.astype(float), 0) #astype float to handle timedeltas. timedelta / timedelta is a float so the final product needs to be a float anyway
    result[mask] = numerator[mask]/denominator[mask]

    ##If option is 1 then return 1 if the numerator and the denominator are the same (both 0 or both inf)
    #todo if useful sign could be included and np.inf / (-np.inf) could calculate to -1
    if option == 1:
        mask = np.isclose(denominator, numerator)
        result[mask] = 1
    return result

def f_bilinear_interpolate(im, x_im, y_im, x, y):
    ##get the index of x and y within the x_im and y_im arrays
    x= np.interp(x, x_im, np.arange(len(x_im)))
    y= np.interp(y, y_im, np.arange(len(y_im)))
    x = np.asarray(x)
    y = np.asarray(y)

    x0 = np.floor(x).astype(int)
    x1 = x0 + 1
    y0 = np.floor(y).astype(int)
    y1 = y0 + 1

    x0 = np.clip(x0, 0, im.shape[1]-1)
    x1 = np.clip(x1, 0, im.shape[1]-1)
    y0 = np.clip(y0, 0, im.shape[0]-1)
    y1 = np.clip(y1, 0, im.shape[0]-1)

    Ia = im[ y0, x0 ]
    Ib = im[ y1, x0 ]
    Ic = im[ y0, x1 ]
    Id = im[ y1, x1 ]

    wa = (x1-x) * (y1-y)
    wb = (x1-x) * (y-y0)
    wc = (x-x0) * (y1-y)
    wd = (x-x0) * (y-y0)

    return wa*Ia + wb*Ib + wc*Ic + wd*Id

def np_extrap(x, xp, yp):
    ## np.interp function with linear extrapolation if x is beyond the input date (xp)
    ### from https://stackoverflow.com/questions/2745329/how-to-make-scipy-interpolate-give-an-extrapolated-result-beyond-the-input-range"""
    x = np.array(x) #convert x to array so that it can be masked
    y = np.array(np.interp(x, xp, yp))  #convert y to array so that it can be masked (required if x is a scalar)
    ##use a mask to adjust values if extrapolating x below the lowest input value in xp
    y[x < xp[0]] = yp[0] + (x[x<xp[0]]-xp[0]) * (yp[0]-yp[1]) / (xp[0]-xp[1])
    ##use a mask to adjust values if extrapolating x above the highest input value in xp
    y[x > xp[-1]]= yp[-1] + (x[x>xp[-1]]-xp[-1])*(yp[-1]-yp[-2])/(xp[-1]-xp[-2])
    return y

def f_norm_cdf(x, mu, cv=0.2, sd=None):
    '''
    ## returns the probability of the value being less than or equal to x
    ## based on a normal distribution with mean mu and either a
    ## coefficient of variation cv (default 20%) or standard deviation sd.
    ## If provided the sd is used.
    ## Calculated using an approximation of the normal probability function
    '''

    ##sd - standard deviation - maximum to stop div0 errors in next step.
    if sd is None:
        sd = mu * cv
    ##standardise x. f_divide in case SD is 0 (either mu is 0 or CV is 0)
    xstd = f_divide(x - mu,  sd)
    ##probability (<=x)
    prob = f_back_transform(358 / 23 * xstd - 111 * np.arctan(37 / 294 * xstd))
#    prob = 1 / (np.exp(-358 / 23 * xstd + 111 * np.arctan(37 / 294 * xstd)) + 1)
    return prob


def f_distribution7(mean, sd=None, cv=None):
    '''
    ##create a distribution around the mean for a variable that can be applied in any non-linear relationships
    ##Create 7 intervals with equal probability
    ## Equal probability allows the non-linear result to be averaged with equal weighting
    '''

    if sd is None:
        sd = cv * mean
    ## The distribution of standardised x based on the mid point of 7 intervals of 14.3%
    dist7_p1 = np.array([-1.535, -0.82, -0.375, 0, 0.375, 0.82, 1.535])
    ## Apply the distribution to the mean using the std deviation
    var_p1 = mean[..., na] + sd[..., na] * dist7_p1
    return var_p1

def f_find_closest(A, target):
    ##info here: https://stackoverflow.com/questions/8914491/finding-the-nearest-value-and-return-the-index-of-array-in-python
    #A must be sorted
    idx = A.searchsorted(target)
    idx = np.clip(idx, 1, len(A)-1)
    left = A[idx-1]
    right = A[idx]
    idx -= target - left < right - target
    return idx

def f_reduce_skipfew(ufunc, foo, preserveAxis=None):
    '''performs function on each axis except the axis that a specified as preserveAxis'''
    r = np.arange(foo.ndim)
    if preserveAxis is not None:
        preserveAxis = tuple(np.delete(r, preserveAxis))
    return ufunc(foo, axis=preserveAxis)

##check if two param dicts are the same.
def findDiff(d1, d2):
    a=False
    ##check if all the keys in d2 exist in d1
    for k in d2:
        if (k not in d1):  # check if the key in current params is in previous params dict.
            # print('DIFFERENT')
            a = True
            return a

    ##checks if all the keys in d1 exist in d2. if so it check the value stored is the same
    for k in d1:
        if a != True: #this stops it looping through the rest of the keys once it finds a difference
            if (k not in d2): #check if the key in current params is in previous params dict.
                # print('DIFFERENT')
                a = True
                return a
            else:
                if type(d1[k]) is dict and type(d2[k]) is dict: #need to check both are dicts in case extra level was added.
                    # print('going level deeper',k)
                    a=findDiff(d1[k],d2[k])
                    # print(k,a)
                else:
                    try: #have to try both ways because sometimes param is array and other times it is scalar
                        if np.any(d1[k] != d2[k]): #if keys are the same, check if the values are the same
                            # print('DIFFERENT',k)
                            a=True
                            return a #only return if true
                    except ValueError: #if the array is a different length then we get value error
                        a = True
                        return a  # only return if true
                    except TypeError:
                        if d1[k] != d2[k]: #if keys are the same, check if the values are the same
                            a=True
                            return a #only return if true
    return a

def f_clean_dict(d):
    '''Replace None values with 0 in a dict.'''
    for k in d:
        if type(d[k]) is dict:  # check if value is a dict. if so go a level deeper
            f_clean_dict(d[k])
        else:
            if d[k] == None:
                d[k] = 0
    return d

def f_dict_reset(used_dict, base_dict):
    '''

    :param used_dict: dictionary that is being reset (must have same keys as base_dict)
    :param base_dict: dictionary with base values used to reset (must have same keys as used_dict).
    :return: None
    '''

    for key in base_dict:
        used_dict[key] = copy.deepcopy(base_dict[key])

def f_produce_df(data, rows, columns, row_names=None, column_names=None):
    """rows is a list of lists that will be used to build a MultiIndex
    columns is a list of lists that will be used to build a MultiIndex"""
    ##if either cols or rows don't exist then add a default 0 as name
    if len(rows) == 0:  #check if no index
        row_index=[0]
    elif not any(isinstance(i, (list, np.ndarray,object)) for i in rows): #check if nested list
        row_index = rows
    elif len(rows)==1: #check if nested list with one element eg don't need to create multiindex
        row_index = rows[0]
    else:
        row_index = pd.MultiIndex.from_product(rows, names=row_names)
    if len(columns) == 0: #check if no index
        col_index=[0]
    elif not any(isinstance(i, (list, np.ndarray,object)) for i in columns): #check if nested list
        col_index = columns
    elif len(columns)==1: #check if nested list with one element eg don't need to create multiindex
        col_index = columns[0]
    else:
        col_index = pd.MultiIndex.from_product(columns, names=column_names)
    return pd.DataFrame(data, index=row_index, columns=col_index)

def f_back_transform(x):
    ''' Back transform a value using a derivation of exp(x) / (1 + exp(x))'''
    return 1 / (1 + np.exp(-x))

def f_sig(x,a,b):
    ''' Sig function CSIRO equation 124 ^the equation below is the sig function from SheepExplorer'''
    return f_back_transform(2 * (np.log(0.95) - np.log(0.05)) / (b-a) * (x - (a+b)/2))
#    return  1/(1+np.exp(-((2*(np.log(0.95) - np.log(0.05))/(b-a))*(x-(a+b)/2))))

def f_ramp(x,a,b):
    ''' RAMP function CSIRO equation 125a'''
    return  np.minimum(1,np.maximum(0,(a-x)/(a-b)))

def f_dim(x,y):
    '''a function that minimum value of zero otherwise difference between the 2 inputs '''
    return np.maximum(0,x-y)

def f_comb(n,k):
    # ##Create an array of factorial values up to n
    # factorial = np.cumprod(np.arange(np.max(n))+1)
    # ##Combination
    # combinations = factorial[n-1]/(factorial[k-1]*factorial[n-k-1])
    ##Create an array of factorial values up to n
    factorial_range = np.arange(np.max(n)+1)
    factorial_range[0] = 1
    factorial = np.cumprod(factorial_range)
    ##Combination
    combinations = factorial[n]/(factorial[k]*factorial[n-k])
    return combinations

def f_dynamic_slice(arr, axis, start, stop, axis2=None, start2=None, stop2=None):
    ##check if arr is int - this is the case for the first loop because arr may be initialised as 0
    if type(arr)==int:
        return arr
    else:
        ##first axis slice if it is not singleton
        if arr.shape[axis]!=1:
            sl = [slice(None)] * arr.ndim
            sl[axis] = slice( start, stop)
            arr = arr[tuple(sl)]
        if axis2 is not None:
            ##second axis slice if required and not singleton
            if arr.shape[axis2] != 1:
                sl = [slice(None)] * arr.ndim
                sl[axis2] = slice( start2, stop2)
                arr = arr[tuple(sl)]
        return arr

def f_nD_interp(x, xp, yp, axis):
    '''
    Interp with multi-D this is essentially the same as looping through axis and applying np.interp

    All inputs must be broadcastable.

    :param x: The x-coordinates at which to evaluate the interpolated values.
    :param xp: The x-coordinates of the data points, must be increasing
    :param yp: The y-coordinates of the data points, same length as xp
    :param axis: Axis to interp along
    :return: y - the interpolated values, same shape as x.
    '''
    ##add new axis where arrays have diff number of dims
    n_dims = max(x.ndim,xp.ndim,yp.ndim)
    extra_axes = tuple(range(n_dims-x.ndim))
    x = np.expand_dims(x,axis=extra_axes)
    extra_axes = tuple(range(n_dims-xp.ndim))
    xp = np.expand_dims(xp,axis=extra_axes)
    extra_axes = tuple(range(n_dims-yp.ndim))
    yp = np.expand_dims(yp,axis=extra_axes)

    ##move axis to interp along into pos=0
    x = np.moveaxis(x, source=axis, destination=0)
    xp = np.moveaxis(xp, source=axis, destination=0)
    yp = np.moveaxis(yp, source=axis, destination=0)
    ##broadcast all arrays to be the same along all axis except the interp axis
    shape = tuple(np.maximum.reduce([xp.shape[1:], yp.shape[1:], x.shape[1:]]))
    x = np.broadcast_to(x, x.shape[0:1]+shape)
    xp = np.broadcast_to(xp, xp.shape[0:1]+shape)
    yp = np.broadcast_to(yp, yp.shape[0:1]+shape)
    ##store shape of final array so it can be reshaped back
    final_shape = x.shape
    ##reshape
    x = x.reshape(x.shape[0],-1)
    xp = xp.reshape(xp.shape[0],-1)
    yp = yp.reshape(yp.shape[0],-1)
    ##loop and do interp
    y = np.zeros(x.shape)
    for i in range(x.shape[-1]):
        y[:,i] = np.interp(x[:,i], xp[:,i], yp[:,i])
    ##reshape to normal
    y = y.reshape(final_shape)
    y = np.moveaxis(y, source=0, destination=axis)
    return y

def f_merge_axis(a, source_axis=0, target_axis=1):
    '''
    This function merges two axis into one. This is basically reshaping but works if the two axis are not side by side.
    :param a: numpy array
    :param source_axis: position of axis being merged with target_axis (this axis wont exist in the new array)
    :param target_axis: position axis to be kept (this is merged with source axis)
    :return:
    '''
    ##determine new shape
    shp = np.array(a.shape)
    shp[target_axis] *= shp[source_axis]
    out_shp = np.delete(shp,source_axis)
    ##If source and target axes are adjacent ones, we can skip moveaxis and simply reshape
    if target_axis==source_axis+1:
        return a.reshape(out_shp)
    else:
        return np.moveaxis(a,source_axis,target_axis-1).reshape(out_shp)

def f_split_axis(a, len_a, axis):
    '''
    This function splits an axis into two axis.

    This is basically reshaping.

    :param a: numpy array
    :param len_a: length of the first of the two new axis
    :return:
    '''
    ##determine new shape
    shp = np.array(a.shape)
    shp[axis] /= len_a
    out_shp = np.insert(shp,axis,len_a)
    return a.reshape(out_shp)


#######################
#Specific AFO function#
#######################
def f_sa(value, sa, sa_type=0, target=0, value_min=-np.inf,pandas=False, axis=0):
    '''applies SA. Function can handle numpy or pandas'''

    ##Type 0 is sam (sensitivity multiplier) - default
    if sa_type == 0:
        if pandas:
            value = np.maximum(value_min, value.mul(sa, axis=axis))
        else:
            value  = np.maximum(value_min, value * sa)
    ##Type 1 is sap (sensitivity proportion, sa = 0 no change, sa = 1 doubles the value)
    elif sa_type == 1:
        if pandas:
            value = np.maximum(value_min, value.mul(1 + sa, axis=axis))
        else:
            value  = np.maximum(value_min, value * (1 + sa))
    ##Type 2 is saa (sensitivity addition)
    elif sa_type == 2:
        try:  #in case value array is datearray and value_min is np.inf
            value = np.maximum(value_min, value + sa)
        except TypeError:
            value = value + sa

    ##Type 3 is sat (sensitivity target, sa=1 returns the target, sa=-1 returns value that is same distance but opposite direction to target)
    elif sa_type == 3:
        if pandas:
            value = np.maximum(value_min, value + (target - value).mul(sa, axis=axis))
        else:
            value  = np.maximum(value_min, value + (target - value) * sa)
    ##Type 4 is sar (sensitivity range. sa=-1 returns 0, sa=1 returns 1)
    elif sa_type == 4:
         value = np.maximum(0, np.minimum(1, value * (1 - np.abs(sa)) + np.maximum(0, sa)))
    ##Type 5 is sav (return the SA value, '-' is no change)
    elif sa_type == 5:
        try:
            sa=sa.copy()#have to copy the np arrays so that the original sa is not changed
        except:
            pass
        value = f_update(value, sa, sa != '-')

    return value


def f_run_required(exp_data1):
    '''
    here we check if precalcs and pyomo need to be recalculated. this is slightly complicated by the fact that columns and rows can be added to exp.xls
    and the fact that a user can opt not to run a trial even if it is out of date so the run requirement must be tracked
    have any sa cols been added or removed, are the values the same, has the py code changed since last run?

    this function is also used by report.py to calculate if reports are being generated with out of date data.
    '''
    ##add run cols to be populated
    exp_data1['run_req'] = False

    ###if only ReportControl.py or ReportFunctions.py have been updated precalcs don't need to be re-run therefore newest is equal to the newest py file that isn't a report
    sorted_list = sorted(glob.iglob('*.py'), key=os.path.getmtime)
    if sorted_list[-1] != 'ReportFunctions.py' and sorted_list[-1] != 'ReportControl.py':
        newest = sorted_list[-1]
    elif sorted_list[-2] != 'ReportFunctions.py' and sorted_list[-2] != 'ReportControl.py':
        newest = sorted_list[-2]
    else:
        newest = sorted_list[-3]

    try: #in case pkl_exp doesn't exist
        with open('pkl/pkl_exp.pkl',"rb") as f:
            prev_exp = pkl.load(f)

        ##get a list of all sa cols (including the name of the trial because two trial may have the same values but have a different name)
        keys_hist = list(prev_exp.reset_index().columns[3:].values)
        keys_current = list(exp_data1.reset_index().columns[3:].values)

        ##update prev_exp run column
        ###if the trial was run the last time the model was run (r_vals are newer than exp.pkl) this trial doesn't need to be re-run unless code or inputs have changed.
        ###if r_vals don't exist the trial needs to be re-run (this allows the user to delete r_vals to re-run a trial).
        run_last = []
        no_r_vals = []
        for trial in prev_exp.index.get_level_values(3):
            try:
                if os.path.getmtime('pkl/pkl_exp.pkl') <= os.path.getmtime('pkl/pkl_r_vals_{0}.pkl'.format(trial)):
                    run_last.append(True)
                    no_r_vals.append(False)
                else:
                    run_last.append(False)
                    no_r_vals.append(False)
            except FileNotFoundError:
                run_last.append(False)
                no_r_vals.append(True)
        prev_exp.loc[run_last, ('run_req', '', '', '')] = False #set run req to false if trial was run last iteration of the model.
        prev_exp.loc[no_r_vals, ('run_req', '', '', '')] = True #set run req to True if r_vals don't exist

        ##if headers are the same, code is the same and the excel inputs are the same then test if the values in exp.xls are the same
        if (keys_current==keys_hist and os.path.getmtime('pkl/pkl_exp.pkl') >= os.path.getmtime(newest)
                                    and os.path.getmtime('pkl/pkl_exp.pkl') >= os.path.getmtime("Universal.xlsx")
                                    and os.path.getmtime('pkl/pkl_exp.pkl') >= os.path.getmtime("Property.xlsx")
                                    and os.path.getmtime('pkl/pkl_exp.pkl') >= os.path.getmtime("Structural.xlsx")):
            ###check if each trial has the same values in exp.xls as last time it was run.
            i3 = prev_exp.reset_index().set_index(keys_hist).index  # have to reset index because the name of the trial is going to be included in the new index so it must first be dropped from current index
            i4 = exp_data1.reset_index().set_index(keys_current).index
            exp_data1.loc[~i4.isin(i3),('run_req', '', '', '')] = True
        ##if headers are different or py code has changed then all trials need to be re-run
        else: exp_data1['run_req']=True
    except FileNotFoundError:
        exp_data1['run_req']=True
    return exp_data1

def f_read_exp():
    '''

    1. Read in exp.xl, set index and cols and drop un-required cols.
    2. Determine which trials are in the experiment the user specified to run.
    '''

    ##set the group of trials being run. If no argument is passed in then all trials are run. To pass in argument need to run via terminal.
    try:
        exp_group = int(sys.argv[1]) #reads in as string so need to convert to int, the script path is the first value hence take the second.
    except IndexError: #in case no arg passed to python
        exp_group = None

    ##read excel
    exp_data = pd.read_excel('exp.xlsx', index_col=None, header=[0,1,2,3], engine='openpyxl')

    ##determine trials which are in specified experiment group. If no group passed in then all trials will be included in the experiment.
    if exp_group is not None:
        exp_group_bool = exp_data.loc[:,('Drop','blank','blank','Exp Group')].values==exp_group
    else:
        exp_group_bool = exp_data.loc[:,('Drop','blank','blank','Exp Group')].values >= 0 #this will remove the blank rows

    ##drop irrelevant cols and set index
    exp_data = exp_data.iloc[:, exp_data.columns.get_level_values(0)!='Drop']
    exp_data = exp_data.set_index(list(exp_data.columns[0:4]))

    ##get the name of each trial in the experiment group
    experiment_trials = exp_data.index.get_level_values(3)[exp_group_bool]

    ##check if any trials have the same name
    if len(experiment_trials) == len(set(experiment_trials)):
        pass
    else:
        raise exc.TrialError('''Exp.xl has multiple trials with the same name.''')

    return exp_data, exp_group_bool


def f_group_exp(exp_data, exp_group_bool):
    '''
    Cuts exp based on the group passed in as argument by user. If no argument then all trials are run.
    This has to be a separate function so that the run required code has access to the full exp.
    '''

    ##cut exp based on group argument
    exp_data = exp_data.loc[exp_group_bool]
    return exp_data

def f_update_sen(row, exp_data, sam, saa, sap, sar, sat, sav, sam_inp, saa_inp, sap_inp, sar_inp, sat_inp, sav_inp):
    ##reset SA dicts to base at the start of each trial before applying SA.
    f_dict_reset(sam, sam_inp)
    f_dict_reset(sap, sap_inp)
    f_dict_reset(saa, saa_inp)
    f_dict_reset(sat, sat_inp)
    f_dict_reset(sar, sar_inp)
    f_dict_reset(sav, sav_inp)

    for dic,key1,key2,indx in exp_data:
        ##extract current value
        value = exp_data.loc[exp_data.index[row], (dic,key1,key2,indx)]

        ##value needs to be single ie don't want a single value series (for some reason sometimes we are getting series)
        if isinstance(value, pd.Series):
            value = value.squeeze()

        ##change indx to str so the following if statements work
        indx = str(indx) #change to string because sometimes blank is read in as nan

        ##checks if both slice and key2 exists
        if not ('Unnamed' in indx  or 'nan' in indx or 'Unnamed' in key2):
            indices = tuple(slice(*(int(i) if i else None for i in part.strip().split(':'))) for part in indx.split(',')) #creats a slice object from a string - note slice objects are not inclusive ie to select the first number it should look like [0:1]
            if dic == 'sam':
                sam[(key1,key2)][indices] = sam[(key1,key2)][indices] * value  # if there are multiple instances of the same SA in exp.xlsx they accumulate
            elif dic == 'sap':
                sap[(key1,key2)][indices] = (1 + sap[(key1,key2)][indices]) * (1 + value) - 1  # if there are multiple instances of the same SA in exp.xlsx they accumulate
            elif dic == 'saa':
                saa[(key1,key2)][indices] = saa[(key1,key2)][indices] + value  # if there are multiple instances of the same SA in exp.xlsx they accumulate
            elif dic == 'sat':
                sat[(key1,key2)][indices] = value   # last entry in exp.xlsx is used
            elif dic == 'sar':
                sar[(key1,key2)][indices] = value   # last entry in exp.xlsx is used, could be changed to accumulate
            elif dic == 'sav':
                if value != "-":
                    sav[(key1,key2)][indices] = value   # last entry in exp.xlsx that is not "-" is used

        ##checks if just slice exists
        elif not ('Unnamed' in indx  or 'nan' in indx):
            indices = tuple(slice(*(int(i) if i else None for i in part.strip().split(':'))) for part in indx.split(',')) #creats a slice object from a string - note slice objects are not inclusive ie to select the first number it should look like [0:1]
            if dic == 'sam':
                sam[key1][indices] = sam[key1][indices] * value  # if there are multiple instances of the same SA in exp.xlsx they accumulate
            elif dic == 'sap':
                sap[key1][indices] = (1 + sap[key1][indices]) * (1 + value) - 1  # if there are multiple instances of the same SA in exp.xlsx they accumulate
            elif dic == 'saa':
                saa[key1][indices] = saa[key1][indices] + value  # if there are multiple instances of the same SA in exp.xlsx they accumulate
            elif dic == 'sat':
                sat[key1][indices] = value
            elif dic == 'sar':
                sar[key1][indices] = value
            elif dic == 'sav':
                if value != "-":
                    sav[key1][indices] = value
        ##checks if just key2 exists
        elif not 'Unnamed' in key2:
            if dic == 'sam':
                sam[(key1,key2)] = sam[(key1,key2)] * value
            elif dic == 'sap':
                sap[(key1,key2)] = (1 + sap[(key1,key2)]) * ( 1+ value) -1
            elif dic == 'saa':
                saa[(key1,key2)] = saa[(key1,key2)] + value
            elif dic == 'sat':
                sat[(key1,key2)] = value
            elif dic == 'sar':
                sar[(key1, key2)] = value
            elif dic == 'sav':
                if value != "-":
                    sav[(key1,key2)] = value
        ##if just key1 exists
        else:
            if dic == 'sam':
                sam[key1] = sam[key1] * value
            elif dic == 'sap':
                sap[key1] = (1 + sap[key1]) * (1 + value) - 1
            elif dic == 'saa':
                saa[key1] = saa[key1] + value
            elif dic == 'sat':
                sat[key1] = value
            elif dic == 'sar':
                sar[key1] = value
            elif dic == 'sav':
                if value != "-":
                    sav[key1] = value

def f1_make_r_val(r_vals, param, name, maskz8=None, z_pos=0, shape=None):
    '''
    This function saves a variable in the r_vals dict so it can be accessed in the reporting stage.

    The majority of this function concerns unclustering the z axis. This is required for two reasons:

        1. By the time the r_val is save it would have likely been masked by mask_z8.
        2. The user may have incorrectly clustered the inputs in excel (eg seasons had different inputs before they
           were identified). This doesn't effect the actual model because z8 is masked until it is identified
           however if the r_val didn't get z8 treatment the reports could contain errors.

    :param r_vals: r_vals dict
    :param param: param to be stored
    :param maskz8: season identification mask
    :param name: name of r_val
    :param shape: optional - shape of r_val - this can be used to remove singleton axis.

    Note 1: Arrays must broadcast.
    Note 2: if no z axis then param is simply stored in r_vals no need to pass in the mask arg.
    '''
    if maskz8 is not None:
        df=False
        series=False
        ##convert df to series
        if isinstance(param,pd.DataFrame):
            df = True
            n_cols = param.columns.nlevels
            param = param.stack(list(range(n_cols)))
        ##convert pd.Series to numpy
        if isinstance(param,pd.Series):
            series = True
            ##store index
            index = param.index
            ##reshape array to be numpy
            reshape_size = param.index.remove_unused_levels().levshape # create a tuple with the rights dimensions
            param = np.reshape(param.values,reshape_size)

        ##uncluster z so that each season gets complete information
        index_z = f_expand(np.arange(maskz8.shape[z_pos]), z_pos)
        a_zcluster = np.maximum.accumulate(index_z * maskz8, axis=z_pos)
        a_zcluster = np.broadcast_to(a_zcluster, param.shape)
        param = np.take_along_axis(param, a_zcluster, axis=z_pos)

        ##add index if pandas
        if series:
            param = pd.Series(param.ravel(), index=index)

        ##unstack back to a df if required
        if df:
            param = param.unstack(list(range(-n_cols,0)))

    ##reshape if required
    if shape is not None:
        param = param.reshape(shape)

    ##store param
    r_vals[name] = param



def f1_make_pyomo_dict(param, index, loop_axis_pos=None, index_loop_axis_pos=None, dtype='float32'):
    '''
    Convert numpy array into dict for pyomo. A loop can be used to reduce memory if required.

    0 values are removed to reduce time (when creating the param in pyomo) and space.

    :param param: numpy array
    :param index: list of index arrays
    :param loop_axis_pos: optional: position of axis that is being looped on (arg not required if no loop)
    :param index_loop_axis_pos: optional: position of axis that is being looped on in the index array (arg not required if no loop)
    :return: dict for pyomo
    '''
    ##build in loop to reduce memory for some big params
    if loop_axis_pos:
        param_masked = np.array([],dtype=dtype)
        index_masked = np.array([])
        for i in range(param.shape[loop_axis_pos]):
            ###mask out values=0
            param_cut = f_dynamic_slice(param, loop_axis_pos, start=i, stop=i+1)
            mask = param_cut != 0
            param_masked = np.concatenate([param_masked,param_cut[mask]],0).astype(dtype)  # applying the mask does the raveling and squeezing of singleton axis
            mask = mask.ravel() #needs to be 1d to mask the index
            ###build index
            ####adjust if the position given is negative (eg cant use pos=-1)
            if index_loop_axis_pos<0:
                index_loop_axis_pos = len(index) + index_loop_axis_pos
            index_cut = [index[x] if x != index_loop_axis_pos else index[x][i:i+1] for x in range(len(index))]
            index_cut = cartesian_product_simple_transpose(index_cut)
            index_masked = np.vstack([index_masked,index_cut[mask,:]]) if index_masked.size else index_cut[mask,:]
    else:
        ###mask out values=0
        mask = param!=0
        ###build index
        index = cartesian_product_simple_transpose(index)
        ###mask param and index
        param_masked = param[mask]  # applying the mask does the raveling and squeezing of array
        mask = mask.ravel() #needs to be 1d to mask the index
        index_masked = index[mask,:]

    ##error check - index and param should be same length but zip() doesn't throw error if they are different length
    if len(index_masked) != len(param_masked):
        raise exc.ParamError('''Index and param must be the same length''')

    ##make index a tuple and zip with param and make dict
    tup = tuple(map(tuple,index_masked))
    return dict(zip(tup, param_masked))

def write_variablesummary(model, row, exp_data, obj, option=0, property_id=''):
    '''

    :param model: pyomo model
    :param row: trial row
    :param exp_data: exp info
    :param obj: objective value
    :param option: 0: trial name and property ID will be included in file name
                   1: file name will be generic
    :return:
    '''
    ##This writes variables with value greater than 0.0001 to txt file
    ### written with trial description in file name if full solution is requested (option 0)
    ### written every iteration with generic name (option 1) - can be used to check progress of analysis each iteration
    directory_path = os.path.dirname(os.path.abspath(__file__))
    if option == 0:
        file = open(os.path.join(directory_path, f'Output/Variable summary {exp_data.index[row][3]} - {property_id}.txt'),'w')  # file name has to have capital
    else:
        file = open(os.path.join(directory_path, 'Output/Variable summary.txt'),'w')  # file name has to have capital
    file.write('Trial: %s\n' % exp_data.index[row][3])  # the first line is the name of the trial
    file.write('{0} profit: {1}\n'.format(exp_data.index[row][3],obj))  # the second line is profit
    for v in model.component_objects(pe.Var,active=True):
        file.write("Variable %s\n" % v)  # \n makes new line
        for index in v:
            try:
                if v[index].value > 0.0001 or v[index].value < -0.0001:
                    file.write("   %s %s\n" % (index,v[index].value))
            except:
                pass
    file.close()



##########################
# period calculators     #
##########################

def period_allocation(period_dates,periods,start_d,length=None):
    '''
    Parameters
    ----------
    period_dates : List
        Dates of the periods you are matching within eg labour periods or cashflow periods
        *note the start date of the period must added to the end of the period if length is passed in
    periods : List
        Name of each period.
    start_d : Date
        Date of interest.
    length : Dt, optional
        Length of the period of interest. The default is ''.

    Returns
    -------
    Proportion of a given date range in each period:
    Either take a date and returns the period it is in
    or take a date and a length and return a dataframe with a proportion in each period

    '''
    #gets the dates
    # period_dates = p_dates   # don't need this step if the variables passed in are changed to period_dates from p_dates and periods from p_name
    #gets the period name
    # periods = p_name
    if length is not None:
    #start empty list to append to
        allocation_period = []
        end = start_d + length
        #check how much of the range falls into each cash period
        for i in range(len(periods)-1):
            ## ^might be simpler to do this with allocation_period.append  \
            ## # (min(per_end,end)-max(per_start,start)/(end-start)) clipped(0,1)
            ## # would also be quicker if the loop started with i = bisect(period_dates,start)-1
            ## # and finished when per_end > end
            ## # perhaps this is a while loop
            per_start = period_dates[i]
            per_end = period_dates[i + 1]

            ##had to add this if statement to handle feed periods - when using fp convert all dates to 2019 but sometimes the start date plus the length = 2020. there for sometimes the end date need to be adjusted back to 2019, when this happens the start date also needs to go back one yr to 2018
            if end -  rdelta.relativedelta(years=1) >= per_start:
                end = (start_d + length) -  rdelta.relativedelta(years=1)
                start = start_d -  rdelta.relativedelta(years=1)
            else:
                end = start_d + length
                start = start_d

            #if the range lasts longer than one cashflow period then that cashflow period gets allocated a proportion
            if start <=  per_start and end >= per_end:
                 allocation_period.append((per_end - per_start) / (end - start))
            #start of the range is before period and the end of the range is after the start of the period but before the end
            elif start <=  per_start <= end and end <= per_end:
                allocation_period.append((end - per_start) / (end - start))
            #is the start of the range after the start of the period and before the finish of the period
            #and the end of the range is after the end of the period
            elif start >=  per_start and start <= per_end <= end :
                allocation_period.append((per_end - start) / (end - start))
            #if all of the range occurs within one the period
            elif start >=  per_start and end <= per_end:
                allocation_period.append(1)
            #if the range doesn't occur in a period.
            else:
                allocation_period.append(np.nan)
        return pd.DataFrame(list(zip(periods,allocation_period)), columns= ('period', 'allocation'))
    #returns the period name a given date falls into
    else:    #^ could use the python function allocation_p = bisect.bisect(period_dates,start)-1
        for date, period_name in zip(period_dates, periods):
            while date <= start_d:
                allocation_p = period_name
                break
        return allocation_p


def f_range_allocation_np(period_dates, item_start, length=np.array([1]), method=1, shape=None):
    ''' Numpy version - The proportion of a date range that falls within each period or proportion of each period that falls in the tested date range.

    Where possible use the default option (method=1). When using method 2 the date range (item start to
    item end) must fall within the period array (the user must make this happen but a warning will be thrown if not).

    Note: if a date range that falls partly in p[0] and partly in p[-1] will get 100% allocated to p[-1].
          This is because we cant have allocation that crosses the season junction. In some cases this may result in over
          allocation to the last period.

    :param period_dates: the start of the periods (including end date of last period). This array must be broadcastable with start
                  (therefore may need to add new axis if start has a dimension). Period axis must be in pos=0.
    :param item_start: the date of the beginning of the date range to test - a numpy array of dates
    :param length: the length of the date range to test - an array of timedelta. Must be broadcastable to start.
    :param method: Controls the proportion calculated. Method 1 returns the proportion of date range in each period.
                     Method 2 returns the proportion of the period in the date range.
    :param shape: this is the shape of returned array, required if both period_dates & start have more than 1 dim

    :return: Numpy array with shape(period_dates, start array). Containing the proportion of the
             respective period for that test date.
    '''
    ##make length at least 1 to stop div 0 if seeding/harv periods are set to 0
    min = np.array([1])
    length = np.maximum(min,length)

    #start empty array to assign to
    if shape==None:
        allocation_period=np.zeros((period_dates.shape[:-1] + item_start.shape),dtype=np.float64)
    else:
        allocation_period=np.zeros(shape,dtype=np.float64)

    ## adjust dates.
    if method==1:
        ##adjust yr of item occurrence
        start_of_periods = period_dates[0,...]
        end_of_periods = start_of_periods + 363 #use 363 because end date is the day before the end otherwise can get item that starts on the last day of periods.
        add_yrs = np.ceil(np.maximum(0,(start_of_periods - item_start) / 364))
        sub_yrs = np.ceil(np.maximum(0,(item_start - end_of_periods) / 364))
        item_start = item_start + add_yrs * 364 - sub_yrs * 364
        ##handle cases where date + length is after the end of periods. in this situation length gets reduced - this is the easiest method and reduces possible errors due to crossing season junction.
        length = np.minimum(length, (period_dates[-1,...] - item_start))
    elif method==2:
        ###adjust the period dates (leave item date the same)
        period_end_dates = period_dates[1:, ...]
        period_start_dates = period_dates[:-1, ...]
        item_end = item_start + length
        add_yrs = np.ceil(np.maximum(0,(item_start - period_end_dates) / 364))
        sub_yrs = np.ceil(np.maximum(0,(period_start_dates - item_end) / 364))
        period_start_dates = period_start_dates + add_yrs * 364 - sub_yrs * 364
        period_end_dates = period_end_dates + add_yrs * 364 - sub_yrs * 364

    ##end of period
    item_end = np.minimum(item_start + length, period_dates[-1]) #minimum ensures that the assigned date range is within the period date range.

    ##checks if user wants the proportion of each period that falls in the tested date range or proportion of date range in each period
    if method==1:
        #check how much of each date range falls within the period (allocate start dates into periods)
        for i in range(len(period_dates)-1):
            per_start = period_dates[i, ...] #[i:i+1] #to keep dim
            per_end = period_dates[i+1, ...]
            calc_start = np.maximum(per_start,item_start)       #select the later of the period start or the start of the range
            calc_end = np.minimum(per_end,item_end)             #select earlier of the period end and the end of the range
            allocation_period[i,...] = np.maximum(0, (calc_end - calc_start) / (item_end - item_start)) #days between calc_end and calc_start (0 if end before start) divided by length of the range
    else:
        ###check date range falls within periods - period adjustment happens above so this should never be an error unless the inputs are bad.
        if not np.all(np.logical_and(np.min(period_start_dates,axis=0) < item_start, item_end < np.max(period_end_dates, axis=0))):
            raise Warning('Trying to allocate periods into a date array but date array range is greater than the period range.'
                          'This indicates poor inputs. You may need to alter the inputs')
        ###check how much of each period falls within the date range (allocate periods into start dates)
        for i in range(len(period_dates)-1):
            per_start = period_start_dates[i, ...] #[i:i+1]
            per_end = period_end_dates[i, ...]
            calc_start = np.maximum(per_start,item_start)       #select the later of the period start or the start of the range
            calc_end = np.minimum(per_end,item_end)            #select earlier of the period end and the end of the range
            allocation_period[i,...] = np.maximum(0, f_divide(calc_end - calc_start
                                                              , per_end - per_start)) #days between calc_end and calc_start (0 if end before start) divided by length of the period, use f_divide in case any period lengths are 0 (this is likely to occur in season version)

    return allocation_period


def period_proportion_np(period_dates, date_array):
    ''' Numpy version - The period that a given date falls in. and the proportion of the way through the period the date occurs.

    Parameters.
    period_dates: The dates of the periods to search within. Must contain the end date of the last period. If multi-D period axis must be pos 0.
    date_array:  The dates to allocate.

    Note: period_dates and date_array must be broadcastable.

    Returns.
    Two Numpy arrays with shape(date_array).
        1 period_array - the period which the values in date_array occur.
        2 proportion_array - how far through the period the date occurs.
    '''

    ##broadcast period_dates so that it has same size axis as date_array - so slicing works
    shape = (period_dates.shape[0],) + date_array.shape
    period_dates = np.broadcast_to(period_dates, shape)

    ##get dtype consistent
    period_dates = period_dates
    date_array = date_array

    ##dates
    dates_start = period_dates[:-1]
    dates_end = period_dates[1:].copy() #so original date array isn't altered when updating year in next step

    ##adjust yr of item occurrence
    start_of_periods = period_dates[0,...]
    end_of_periods = period_dates[-1,...]
    add_yrs = np.ceil(np.maximum(0,(start_of_periods - date_array) / 364))
    sub_yrs = np.ceil(np.maximum(0,(date_array - end_of_periods) / 364))
    date_array = date_array + add_yrs * 364 - sub_yrs * 364
    ###little check to ensure that all cashflow is all starting at least 1 day before the end cashflow date
    date_array = date_array - np.maximum(0, (date_array - (period_dates[-1,...] - 1)))

    ##calc the period each value in the date array falls within (can't use np.searchsorted because date array has z axis)
    ###occur is bool array which is true for the period that the date array fall into
    occur = np.logical_and(dates_start <= date_array, date_array < dates_end)
    ###period index
    p_idx = np.arange(period_dates[:-1].shape[0])
    ###mul occur and idx to return the period number the date falls in else a 0. then sum the period axis to return the period array
    occur = np.moveaxis(occur,0,-1) #so that period axis is at end
    period_array = np.sum(occur * p_idx, axis=-1)

    ##calc proportion
    per_start = np.take_along_axis(period_dates,period_array[None,...],0)[0]
    per_end = np.take_along_axis(period_dates,period_array[None,...]+1,0)[0]
    # per_start = period_dates[period_array, np.arange(date_array.shape[0])[:,None], np.arange(date_array.shape[1])] #problem is that this is fixed to 3d
    # per_end   = period_array[period_array + 1]
    proportion_array = (date_array - per_start) / (per_end - per_start)
    # print('propn, date, stat, end, start', proportion_array,date_array,per_start,per_end,per_start)
    return period_array, proportion_array

##################
#timing functions#
##################

def f_daylength(dayOfYear, lat):
    """Computes the length of the day (the time between sunrise and
    sunset) given the day of the year and latitude of the location.
    Function uses the Brock model for the computations.
    For more information see, for example,
    Forsythe et al., "A model comparison for daylength as a
    function of latitude and day of year", Ecological Modelling,
    1995.
    Parameters
    ----------
    dayOfYear : int
        The day of the year. 1 corresponds to 1st of January
        and 365 to 31st December (on a non-leap year).
    lat : float
        Latitude of the location in degrees. Positive values
        for north and negative for south.
    Returns
    -------
    d : float
        Daylength in hours.
    """
    dl=np.zeros_like(dayOfYear, dtype='float64')
    latInRad = np.deg2rad(lat)
    declinationOfEarth = 23.45*np.sin(np.deg2rad(360.0*(283.0+dayOfYear)/365.0))
    p1 = (-np.tan(latInRad) * np.tan(np.deg2rad(declinationOfEarth))) <= -1.0
    p2 = (-np.tan(latInRad) * np.tan(np.deg2rad(declinationOfEarth))) >= 1.0
    hourAngle = np.rad2deg(np.arccos(-np.tan(latInRad) * np.tan(np.deg2rad(declinationOfEarth))))
    daylen = 2.0*hourAngle/15.0
    dl[p1] = 24
    dl[p2] = 0
    #todo Should this be logical_or? JMY thinks it should be.
    dl[~np.logical_and(p2, p1)] = daylen[~np.logical_and(p2, p1)]
    return dl


def f_next_prev_association(datearray_slice,*args):
    '''
    Depending on the inputs this function will return the next or previous association.
    eg it can be used to determine the next lambing opportunity for each period.
    See john stuff.py for alternative methods.

    Parameters
    ----------
    datearray_slice : any
        This is 1d array which the second array is sorted into (this must be sorted).
    *args : 1 - array, 2 - int
        Arg 1: the period array 1d that is the index is being found for, note the index is based off the start date therefore must do [1:-1] if you want idx based on end date.
        Arg 2: period offset (this may be needed if evaluating the end of the period.

    Returns
    -------
    Array
        The function finds the index value of the datearray which is either the next or previous date for a given input date.
        ## The previous opportunity is the latest opportunity date that is less than the date at the end of the period
        ## eg ('end of the period' so that if joining occurs during the period it is the previous
        ## The next opportunity is the earliest joining date that is greater than the date at the start of the period
        ## eg So it is the prev + 1 except if the joining is occurring within the period, in which case it points to this one.


    '''
    date=args[0]
    offset=args[1] #offset is used to get the previous datearray period
    side=args[2]
    idx_next = np.searchsorted(datearray_slice, date, side)
    idx = np.clip(idx_next - offset, 0, len(datearray_slice)-1) #makes the max value equal to the length of joining array, because if the period date is after the last lambing opportunity there is no 'next'
    return idx

