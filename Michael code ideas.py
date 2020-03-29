# -*- coding: utf-8 -*-
"""
Created on Sat Jan  4 17:39:41 2020

@author: young
"""
#################################################################################################################################################################
#################################################################################################################################################################
#actual code
#################################################################################################################################################################
#################################################################################################################################################################

###########################
#alternative code methods #
###########################

# 1a - disagregated version of rotation constraint, was used to help debug
def rot_phase_link(model,l,h1,h2,h3,h4):
    print('history :',((h1,)+(h2,)+(h3,)+(h4,)))
    test=0
    for r in model.phases:
        index=((r)+(h1,)+(h2,)+(h3,)+(h4,))
        #print('initial: ',index)
        if index in model.rot_phase_link:
            print('success: ',index)
            test+=model.num_phase[r,l]*model.rot_phase_link[r,h1,h2,h3,h4]
            # print('success to follow')
    print(test)
    return  test<=0
model.jj = Constraint(model.lmus,model.rot_constraints, rule=rot_phase_link, doc='rotation phases constraint')

## another example of dis agregated constraint
def rotation_yield_transfer2(model,k):
    i=uinp.structure['phase_len']-1
    jj=0
    while jj == 0:
        for l in model.s_lmus:
            for h, r in zip(model.s_phases_dis, model.s_phases):
                print(k,h,r)
                print(h[i])
                if ((h[0:i])+(k,)+(l,)) in model.p_rotation_yield:
                    print('first')
                    if model.p_rotation_yield[h[0:i],k,l] != 0 :
                        print('second')
                        if h[i]==k:
                            jj=2
                            print('third')
                            print(model.p_rotation_yield[h[0:i],k,l],model.v_phase_area[r,l])
                        else: print('no')

##Run whole module - runs from here                                                                              
# processes = ('Exp.py', 'Exp.py')                                    
# def run_process(process):                                                             
#     os.system('python {}'.format(process))                                       
                                                                                
# def main():                                                                               
#     pool = Pool(processes=2)
#     # run_process('Exp.py') 
#     pool.map(run_process, processes) 
#     print('Done')                                                        

# if __name__ == '__main__':                                                                               
#     main()


##run just the function - has to be done through anaconda prompt
def main():
    p = multiprocessing.Process(target=exp.exp, args=(0,))
    p1 = multiprocessing.Process(target=exp.exp, args=(1,))
    p2 = multiprocessing.Process(target=exp.exp, args=(1,))
    p3 = multiprocessing.Process(target=exp.exp, args=(1,))
    p4 = multiprocessing.Process(target=exp.exp, args=(1,))
    p5 = multiprocessing.Process(target=exp.exp, args=(1,))
    p.start()
    p1.start()
    p2.start()
    p3.start()
    p4.start()
    p5.start()
    print('done')
    # p2 = multiprocessing.Process(target=hang, args=("2"))
    # p2.start()
    # p.join()
    # p1.join()
    # p2.join()

if __name__ == '__main__':
    start = time.time()
    main()
    end = time.time()
    print('total time',end-start)
######################################
#Write default SA to row1 in exp.xlsx#
######################################
##may want to turn this into a function and only call when you want because it might be slow
import openpyxl
myworkbook=openpyxl.load_workbook('exp1.xlsx')
ws= myworkbook['Sheet1']
c=2
r=5
##checks if the column has a sa value
while not ws.cell(row=1,column=c).value == None:
    dic=ws.cell(row=1,column=c).value
    key1=ws.cell(row=2,column=c).value
    key2=ws.cell(row=3,column=c).value
    indx=ws.cell(row=4,column=c).value
    ##checks if both slice and key2 exists
    if not (indx == None  and key2== None):
        indices = tuple(slice(*(int(i) if i else None for i in part.strip().split(':'))) for part in indx.split(',')) #creats a slice object from a string - note slice objects are not inclusive ie to select the first number it should look like [0:1]
        if dic == 'sam':
            value=sen.sam[(key1,key2)][indices]
        elif dic == 'saa':
            value=sen.saa[(key1,key2)][indices]
        elif dic == 'sap':
            value=sen.sap[(key1,key2)][indices]

    ##checks if just slice exists
    elif not indx == None:
        indices = tuple(slice(*(int(i) if i else None for i in part.strip().split(':'))) for part in indx.split(',')) #creats a slice object from a string - note slice objects are not inclusive ie to select the first number it should look like [0:1]
        if dic == 'sam':
            value=sen.sam[key1][indices]
        elif dic == 'saa':
            value=sen.saa[key1][indices]
        elif dic == 'sap':
            value=sen.sap[key1][indices]
    ##checks if just key2 exists
    elif not key2 == None:
        if dic == 'sam':
            value=sen.sam[(key1,key2)]
        elif dic == 'saa':
            value=sen.saa[(key1,key2)]
        elif dic == 'sap':
            value=sen.sap[(key1,key2)]
    default=value
    try:
        ws.cell(row=r,column=c).value = default
    except ValueError:
        ws.cell(row=r,column=c).value = default[0] #this is needed if you are assigning one value to multiple spots in an array
    c+=1
##it wont save for some reason, it should work but it won't
myworkbook.save('exp1.xlsx')

# ##can use this method as temp sol but the formatting is lost
# import xlsxwriter
# workbook = xlsxwriter.Workbook('hello.xlsx') 
# worksheet = workbook.add_worksheet() 
# for row in ws:
#     for cell in row:
#         ws2[cell.coordinate].value = cell.value
#         worksheet.write(cell.coordinate, cell.value) 
# workbook.close() 



############################################
#previous code methods - depreciated code  #
############################################

#origional method for determining phases df
phases_df = fun.phases(inp.input_data['phases'],inp.input_data['phase_len'])




#################################################################################################################################################################
#################################################################################################################################################################
# general coding info 
#################################################################################################################################################################
#################################################################################################################################################################
    ##calling a function multiple times takes time. call it once and assign result to a unique variable. 
    ##local variables are easier for pyton to locate

############################
#df                        #
############################

'''
Change index - this is useful when you want to preform calcs with multiple dfs
    - reindex, this is easy when there is one common level incommon, can also be done when nothing incommon but often results in nan
    -pd.concat, keys - can be used to add level to df (ie in stubble module; pd.concat([vol]*len(inp.input_data['sheep_pools']), keys=inp.input_data['sheep_pools']))
    -df.columns = pd.MultiIndex.from_product([df.columns, ['C']]) this adds c to the index, so you would then have to use reindex to apply (ie grain price in crop module)

Df to Series
    -.iloc[:,0]
    -.squeeze()

Index with multi index
    -weights=data.loc[:,(slice(None),'Weight')] you can use slice(None) when you want all on one level 

Return a list of column names for a given level
    - list(weights.columns.levels[0])

#delete col with no header #
    this is here becayse dropping a col with no header is not compatible with the usual way
    this if statement checks if the column has a header and then deletes the nessecary way
        if list(df.columns)[0]:
            df=df.drop(df.columns[0],axis=1)
            
Multiply
    -can use np.multiply(df1,df2) if you don't want to worry about headers and indexes being the same

Add new col to df without getting settingwithcopy warning
    - df1.loc[:,'e'] = pd.Series(np.random.randn(sLength), index=df1.index)
    - df1 = df1.assign(e=pd.Series(np.random.randn(sLength)).values)


'''

############################
#dicts                     #
############################
''' 
df to dict;
    - default; colunm names become keys. but colums of the df must be selected to stop it becoming a nested dict (using .to_dict())
    - can transpose to get index as key and cols as values eg:
        mmmm = df.transpose()['Col_1'].to_dict() #uses index as key and col as values
    - however you if you stack the df so that it is 1D then the index (row) will become the keys (remember to set the index before stacking)
    - you can also select two columns one with the keys one with the value using the dict(args) function
            eg if lakes is your DataFrame, you can do something like: 
                area_dict = dict(zip(lakes.area, lakes.count)) #takes two cols, the first becomes key second value
'''

############################
#time series               #
############################

''' 
converrt to int number of days
    - int.days() 

simple time conversion (python basically guesses what time you mean from inputs) -good for simple dates ie 1-1-19
    - date = parse(key, dayfirst = True)   
for more complex ones you can use sharfttime or something like that

Convert dt to int
    -To convert to exact days ie 1day 12hours would convert to 1.5 (D can be changed to M,Y,s for other conversions)
        -date / np.timedelta64(1, 'D')
    - Another method to convert however this doesn't do decimals ie 1day 12hours would convert to 1. To get around this you can convert to seconds then divide by 86400 (60*60*24) to get days as int 
        -.astype('timedelta64[s]') - can use D to get days
    
Access an attribute (ie get the number of days in a dt)
    - Use the dt.days attribute. Access this attribute via: (note if it is 1day 0 seconds, dt.seconds will return 0 not 3600)
        -.dt.days (You can also get the seconds and microseconds attributes in the same way)
        
convert timestamp to datetime - example in range function
    - .date() 
convert data type- example in range function
    - astype(.astype('datetime64[D]'))
'''

############################
#pyomo                     #
############################
'''
Pyomo - if there is a complex variable bound ie one that may be different for different variables within a set such as casual labour which may
        have a different bound for each labour period, you can use constraint to bound it because the variable bounds are not flexible enough
        
delete blocks - sometimes the index also needs to be deleted (not sure what that really means or why it is required)
    -model.del_component(model.num_phase_index)
    -model.del_component(model.num_phase)

Summing with constraints:
    -only sum around parts of the constrain that contain the given set ie sum(rotationcost[r,l]*v_num_rotations[r,l]+othercost[r] for r in rotations)
    -if you sum around something that doesn't have the given set it will use the parameter for each of the loops ie sum(rotationcost[r,l]*v_num_rotations[r,l]+othercost[l] for r in rotations) this will add the other cost for each rotation which may not be wanted

Solver error
ERROR: Solver (glpk) returned non-zero return code (1)
ERROR: See the solver log above for diagnostic information.
-this may be due to an inf number as a param

 
Duals/RC/Slacks  
 ##code below will access duals on constraint
    for c in model.component_objects(pe.Constraint, active=True):
        print ("   Constraint",c)
        for index in c:
            print ("      ", index, model.dual[c[index]])

    ##code below will access slacks on constraint
    for c in model.component_objects(pe.Constraint, active=True):
        print ("   Constraint",c)
        for index in c:
            try:
                print ("      ", index, c[index].uslack())
            except: pass
    ##code below will access rc on variables
    for v in model.component_objects(pe.Var, active=True):
        print ("   Constraint",v)
        for index in v:
            print ("      ", index, model.rc[v[index]])
'''


############################
#time code                 #
############################
'''
import timeit
print(timeit.timeit(test,number=10)/10)
'''

############################
#write to existing excel   #
############################
'''
using the standart write method to exvel as used in stybble sim and rotation gen overwrites any existing sheets (even if they have a diff name)
the method below is a way around, but the workbook must exist already
    book = load_workbook('Rotation.xlsx')
    writer = pd.ExcelWriter('Rotation.xlsx', engine='openpyxl') 
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    mps_bool.to_excel(writer, 'rotation_con1',index=True,header=False)
'''

#################################################################################################################################################################
#################################################################################################################################################################
## possible causes of error#
#################################################################################################################################################################
#################################################################################################################################################################

#casual and perm staff now require farmer to supervise, this could be a somethig that makes the model behave strange (casual staff could become limited due to lack of supervision time).
